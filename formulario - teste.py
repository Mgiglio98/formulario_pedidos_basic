import streamlit as st
from datetime import date
from openpyxl import load_workbook
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import smtplib
from io import BytesIO
import sqlite3, json
from datetime import datetime

# --- Inicializa banco SQLite ---
def init_db():
    conn = sqlite3.connect("pedidos.db")
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS pedidos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            oc TEXT,
            pedido_numero TEXT,
            solicitante TEXT,
            executivo TEXT,
            obra TEXT,
            cnpj TEXT,
            endereco TEXT,
            cep TEXT,
            data_pedido TEXT,
            insumos TEXT,
            data_envio TEXT
        )
    """)
    conn.commit()
    conn.close()

init_db()

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Pedido de Materiais", page_icon="📦")

# Exemplo de credenciais (altere/expanda conforme necessário)
# Chave = login da obra (ex.: "OC2212"), Valor = senha
VALID_LOGINS = {
    "OC2212": "Osborne",
    # adicione outras OCs aqui: "OC1234": "Osborne",
}

def autenticar(login: str, senha: str) -> bool:
    """Retorna True se login/senha conferem (case-sensitive na senha)."""
    if not login or not senha:
        return False
    senha_certa = VALID_LOGINS.get(login.strip())
    return senha_certa is not None and senha.strip() == senha_certa

# Inicializa estado de sessão de login
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False
if "login_oc" not in st.session_state:
    st.session_state.login_oc = None
if "login_time" not in st.session_state:
    st.session_state.login_time = None

def render_login():
    st.markdown("## 🔐 Acesso à obra")
    st.caption("Entre com o login e senha fornecidos para a sua obra.")
    col_a, col_b = st.columns([1,1])
    with col_a:
        login_input = st.text_input("Login (ex.: OC2212)", key="__login_input")
    with col_b:
        senha_input = st.text_input("Senha", type="password", key="__senha_input")

    c1, c2, c3 = st.columns([1,1,2])
    with c1:
        entrar = st.button("Entrar", use_container_width=True)
    with c2:
        limpar = st.button("Limpar", use_container_width=True)

    if limpar:
        st.session_state["__login_input"] = ""
        st.session_state["__senha_input"] = ""
        st.rerun()

    if entrar:
        if autenticar(st.session_state["__login_input"], st.session_state["__senha_input"]):
            st.session_state.logged_in = True
            st.session_state.login_oc = st.session_state["__login_input"].strip()
            st.session_state.login_time = datetime.now()
            # opcional: zere estados antigos do formulário ao logar
            for k in ["insumos", "excel_bytes", "nome_arquivo", "pedido_numero",
                      "solicitante", "executivo", "obra_selecionada", "cnpj",
                      "endereco", "cep", "data_pedido"]:
                if k in st.session_state:
                    try:
                        del st.session_state[k]
                    except:
                        pass
            st.rerun()
        else:
            st.error("Login ou senha inválidos. Verifique e tente novamente.")

# 🔒 Gate: se não logado, mostra tela de login e interrompe o app
if not st.session_state.logged_in:
    render_login()
    st.stop()

# (Opcional) Barra de status do login + botão sair
st.sidebar.markdown(f"**OC logada:** {st.session_state.login_oc}")
if st.sidebar.button("Sair"):
    for k in list(st.session_state.keys()):
        del st.session_state[k]
    st.rerun()

def carregar_ultimo_pedido(oc):
    conn = sqlite3.connect("pedidos.db")
    c = conn.cursor()
    c.execute("SELECT * FROM pedidos WHERE oc = ? ORDER BY id DESC LIMIT 1", (oc,))
    row = c.fetchone()
    conn.close()
    if row:
        # Preenche os campos automaticamente
        (_, _, pedido_numero, solicitante, executivo, obra,
         cnpj, endereco, cep, data_pedido, insumos_json, _) = row
        st.session_state.pedido_numero = pedido_numero or ""
        st.session_state.solicitante = solicitante or ""
        st.session_state.executivo = executivo or ""
        st.session_state.obra_selecionada = obra or ""
        st.session_state.cnpj = cnpj or ""
        st.session_state.endereco = endereco or ""
        st.session_state.cep = cep or ""
        st.session_state.data_pedido = datetime.strptime(data_pedido, "%Y-%m-%d").date()
        st.session_state.insumos = json.loads(insumos_json)
        return True
    return False

# --- CSS (espaçamento) ---
st.markdown("""
<style>
[data-testid="stAppViewContainer"] .main .block-container {
    padding-top: 0.5rem;
    padding-bottom: 2rem;
}
</style>
""", unsafe_allow_html=True)

# --- INICIALIZAÇÃO DE SESSÃO ---
if "insumos" not in st.session_state:
    st.session_state.insumos = []
if "resetar_insumo" not in st.session_state:
    st.session_state.resetar_insumo = False
if "resetar_pedido" not in st.session_state:
    st.session_state.resetar_pedido = False
if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None
if "nome_arquivo" not in st.session_state:
    st.session_state.nome_arquivo = ""

# --- RERUN APÓS DOWNLOAD ---
if st.session_state.get("rerun_depois_download", False):
    st.session_state.rerun_depois_download = False
    for campo in [
        "pedido_numero", "solicitante", "executivo", "obra_selecionada",
        "cnpj", "endereco", "cep", "data_pedido", "excel_bytes",
        "nome_arquivo", "pedido_enviado"
    ]:
        if campo in st.session_state:
            del st.session_state[campo]
    st.session_state.insumos = []
    st.rerun()

# --- CAMPOS PADRÃO ---
for campo in ["pedido_numero", "solicitante", "executivo", "obra_selecionada", "cnpj", "endereco", "cep"]:
    if campo not in st.session_state:
        st.session_state[campo] = ""

if "data_pedido" not in st.session_state:
    st.session_state.data_pedido = date.today()

# --- FUNÇÕES AUXILIARES ---
def enviar_email_pedido(assunto, arquivo_bytes, insumos_adicionados, df_insumos):
    """Envia o e-mail do pedido e, se houver, alerta de insumos sem código."""
    smtp_server = "smtp.office365.com"
    smtp_port = 587
    smtp_user = "matheus.almeida@osborne.com.br"
    smtp_password = st.secrets["SMTP_PASSWORD"]

    # --- Endereços de cópia (fixos + administrativo da obra) ---
    cc_addr = []
    
    # Adiciona o administrativo da obra, se existir
    adm_email = ADM_EMAILS.get(st.session_state.get("adm_obra"))
    if adm_email and adm_email not in cc_addr:
        cc_addr.append(adm_email)

    # --- Classificação dos insumos ---
    basicos, especificos, sem_codigo = [], [], []

    for item in insumos_adicionados:
        qtd = item["quantidade"]
        descricao = item["descricao"]
        codigo = item.get("codigo", "")

        if not codigo or str(codigo).strip() == "":
            sem_codigo.append(f"{descricao} — {qtd}")
            continue

        linha_df = df_insumos[df_insumos["Descrição"] == item["descricao"]]
        if not linha_df.empty and linha_df.iloc[0]["Basico"]:
            max_qtd = linha_df.iloc[0]["Max"]
            if pd.notna(max_qtd) and qtd <= max_qtd:
                basicos.append(f"{item['descricao']} — {qtd}")
            else:
                especificos.append(f"{item['descricao']} — {qtd}")
        else:
            especificos.append(f"{item['descricao']} — {qtd}")

    # --- E-mail principal ---
    corpo_principal = (
        "✅ Novo pedido recebido!\n\n"
        "📄 Materiais Básicos:\n" + ("\n".join(basicos) if basicos else "Nenhum") +
        "\n\n🛠️ Materiais Específicos:\n" + ("\n".join(especificos) if especificos else "Nenhum") +
        "\n\n📌 Insumos sem código cadastrado:\n" + ("\n".join(sem_codigo) if sem_codigo else "Nenhum")
    )

    msg = MIMEMultipart()
    msg["From"] = smtp_user
    msg["To"] = smtp_user
    msg["Cc"] = ", ".join(cc_addr)
    msg["Subject"] = assunto
    msg.attach(MIMEText(corpo_principal, "plain"))

    anexo = MIMEApplication(
        arquivo_bytes,
        _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    anexo.add_header('Content-Disposition', 'attachment', filename="Pedido.xlsx")
    msg.attach(anexo)

    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_password)
        server.send_message(msg)
        print("📨 E-mail principal enviado com sucesso!")
    except Exception as e:
        print(f"Erro ao enviar e-mail principal: {e}")
        return

    # --- E-mail auxiliar para insumos sem código ---
    if sem_codigo:
        try:
            msg_aux = MIMEMultipart()
            msg_aux["From"] = smtp_user
            msg_aux["To"] = smtp_user
            msg_aux["Cc"] = ", ".join(cc_addr)
            msg_aux["Subject"] = f"[Verificação de Insumos] {assunto}"

            corpo_aux = (
                "Olá!\n\n"
                f"Foi recebido no pedido {assunto} os seguintes insumos sem código cadastrado:\n\n"
                + "\n".join(sem_codigo) +
                "\n\nConsegue verificar, por favor, se eles já estão cadastrados no sistema?\n"
                "Se sim, poderia informar o código correto de cada um?\n"
                "Se não, favor realizar a inclusão e me confirmar aqui.\n\n"
                "Obrigado!"
            )

            msg_aux.attach(MIMEText(corpo_aux, "plain"))
            server.send_message(msg_aux)
            print("📩 E-mail auxiliar de verificação enviado com sucesso!")
        except Exception as e:
            print(f"Erro ao enviar e-mail auxiliar: {e}")
        finally:
            server.quit()
    else:
        server.quit()

def carregar_dados():
    """Carrega dados de empreendimentos e insumos."""
    df_empreend = pd.read_excel("Empreendimentos.xlsx")
    df_empreend.columns = df_empreend.columns.str.strip().str.upper()

    df_insumos = pd.read_excel("Insumos.xlsx")
    df_insumos["Min"] = pd.to_numeric(df_insumos.iloc[:, 3], errors="coerce")
    df_insumos["Max"] = pd.to_numeric(df_insumos.iloc[:, 4], errors="coerce")
    df_insumos["Basico"] = df_insumos["Min"].notna() & df_insumos["Max"].notna()
    df_insumos = df_insumos[df_insumos["Descrição"].notna() & (df_insumos["Descrição"].str.strip() != "")]

    df_empreend.loc[-1] = [""] * df_empreend.shape[1]
    df_empreend.index = df_empreend.index + 1
    df_empreend = df_empreend.sort_index()

    insumos_vazios = pd.DataFrame({"Código": [""], "Descrição": [""], "Unidade": [""]})
    df_insumos = pd.concat([insumos_vazios, df_insumos], ignore_index=True)
    return df_empreend, df_insumos

# --- CARREGAMENTO DE DADOS ---
df_empreend, df_insumos = carregar_dados()

# --- LIMPEZA APÓS ENVIO ---
if st.session_state.get("limpar_pedido", False):
    for campo in ["pedido_numero", "solicitante", "executivo", "obra_selecionada", "cnpj", "endereco", "cep"]:
        if campo in st.session_state:
            try:
                st.session_state[campo] = ""
            except Exception:
                pass
    st.session_state.data_pedido = date.today()
    st.session_state.insumos = []
    st.session_state.limpar_pedido = False

# --- LOGO E CABEÇALHO ---
col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    st.image("logo.png", width=300)

st.markdown("""
<div style='text-align: center;'>
    <h2 style='color: #000000;'>Pedido de Materiais</h2>
    <p style='font-size: 14px; color: #555;'>
        Preencha os campos com atenção. Verifique se todos os dados estão corretos antes de enviar.<br>
        Ao finalizar, o pedido será automaticamente enviado para o e-mail do setor de Suprimentos.<br>
        Você poderá baixar a planilha gerada após o envio, para registro ou controle.
    </p>
</div>
""", unsafe_allow_html=True)

# --- BOTÃO PARA CARREGAR ÚLTIMO PEDIDO ---
st.markdown("### 🔄 Reabrir Pedido Anterior")
if st.button("📝 Carregar último pedido"):
    if carregar_ultimo_pedido(st.session_state.login_oc):
        st.success("✅ Último pedido carregado com sucesso! Você pode editá-lo e reenviar.")
    else:
        st.info("ℹ️ Nenhum pedido anterior encontrado para esta obra.")

st.divider()

# --- DADOS DO PEDIDO ---
with st.expander("📋 Dados do Pedido", expanded=True):
    if st.session_state.resetar_pedido:
        st.session_state.pedido_numero = ""
        st.session_state.data_pedido = date.today()
        st.session_state.solicitante = ""
        st.session_state.executivo = ""
        st.session_state.adm_obra = ""
        st.session_state.obra_selecionada = ""
        st.session_state.cnpj = ""
        st.session_state.endereco = ""
        st.session_state.cep = ""
        st.session_state.resetar_pedido = False

    col1, col2 = st.columns(2)
    with col1:
        pedido_numero = st.text_input("Pedido Nº", key="pedido_numero")
        solicitante = st.text_input("Solicitante", key="solicitante")
        obra_selecionada = st.selectbox("Obra", df_empreend["EMPREENDIMENTO"].unique(), index=0, key="obra_selecionada")
    with col2:
        data_pedido = st.date_input(
            "Data",
            key="data_pedido",
            value=st.session_state.data_pedido if "data_pedido" in st.session_state else date.today()
        )
        executivo = st.text_input("Executivo", key="executivo")
        
        # --- Novo campo: Administrativo da Obra ---
        ADM_EMAILS = {
            "Maria Eduarda": "maria.eduarda@osborne.com.br",
            "Joice": "joice.oliveira@osborne.com.br",
            "Micaele": "micaele.ferreira@osborne.com.br",
            "Graziele": "graziele.horacio@osborne.com.br",
            "Fabio": "fabio.maia@osborne.com.br",
            "Roberto": "roberto.santos@osborne.com.br"
        }
        
        opcoes_adm = [""] + list(ADM_EMAILS.keys())  # primeira opção em branco
        adm_obra = st.selectbox(
            "Administrativo da Obra",
            opcoes_adm,
            index=0,
            key="adm_obra"
        )

    if obra_selecionada:
        dados_obra = df_empreend[df_empreend["EMPREENDIMENTO"] == obra_selecionada].iloc[0]
        st.session_state.cnpj = dados_obra["CNPJ"]
        st.session_state.endereco = dados_obra["ENDERECO"]
        st.session_state.cep = dados_obra["CEP"]

    st.text_input("CNPJ/CPF", value=st.session_state.get("cnpj", ""), disabled=True)
    st.text_input("Endereço", value=st.session_state.get("endereco", ""), disabled=True)
    st.text_input("CEP", value=st.session_state.get("cep", ""), disabled=True)

st.divider()

# --- ADIÇÃO DE INSUMOS ---
with st.expander("➕ Adicionar Insumo", expanded=True):
    df_insumos_lista = df_insumos.sort_values(by="Descrição", ascending=True).copy()
    df_insumos_lista["opcao_exibicao"] = df_insumos_lista.apply(
        lambda x: f"{x['Descrição']} – {x['Código']} ({x['Unidade']})" if pd.notna(x["Código"]) and str(x["Código"]).strip() != "" else x["Descrição"],
        axis=1
    )

    descricao_exibicao = st.selectbox(
        "Descrição do insumo (Digite em MAIÚSCULO)",
        df_insumos_lista["opcao_exibicao"],
        key="descricao_exibicao"
    )

    dados_insumo = df_insumos_lista[df_insumos_lista["opcao_exibicao"] == descricao_exibicao].iloc[0]
    usando_base = bool(dados_insumo["Código"]) and str(dados_insumo["Código"]).strip() != ""

    if usando_base:
        st.session_state.codigo = dados_insumo["Código"]
        st.session_state.unidade = dados_insumo["Unidade"]
        st.session_state.descricao = dados_insumo["Descrição"]
    else:
        st.session_state.codigo = ""
        st.session_state.descricao = ""

    if "unidade" not in st.session_state or not st.session_state.unidade:
        st.session_state.unidade = ""

    st.write("Ou preencha manualmente o Nome e Unidade se não estiver listado:")

    descricao_livre = st.text_input("Nome do insumo (livre)", key="descricao_livre", disabled=usando_base)
    st.text_input("Código do insumo", key="codigo", disabled=True)
    st.text_input("Unidade", key="unidade", disabled=usando_base)
    quantidade = st.number_input("Quantidade", min_value=1, step=1, format="%d", key="quantidade")
    complemento = st.text_area(
        "Complemento, se necessário (Utilize para especificar medidas, marcas, cores e/ou tamanhos)",
        key="complemento"
    )

    if st.button("➕ Adicionar insumo"):
        descricao_final = st.session_state.descricao if usando_base else descricao_livre

        if descricao_final and quantidade > 0 and (usando_base or st.session_state.unidade.strip()):
            novo_insumo = {
                "descricao": descricao_final,
                "codigo": st.session_state.codigo if usando_base else "",
                "unidade": st.session_state.unidade,
                "quantidade": quantidade,
                "complemento": complemento,
            }
        
            # Se estiver editando um item existente, substitui ele
            if "editando_insumo" in st.session_state and st.session_state.editando_insumo is not None:
                st.session_state.insumos[st.session_state.editando_insumo] = novo_insumo
                st.session_state.editando_insumo = None
                st.success("✏️ Insumo atualizado com sucesso!")
            else:
                st.session_state.insumos.append(novo_insumo)
                st.success("✅ Insumo adicionado com sucesso!")
        
            # Limpa os campos após adicionar/editar
            for campo in ["descricao_exibicao", "descricao_livre", "codigo", "unidade", "quantidade", "complemento"]:
                if campo in st.session_state:
                    try:
                        del st.session_state[campo]
                    except Exception:
                        pass
        
            st.session_state.quantidade = 1
            st.session_state.descricao_exibicao = df_insumos_lista["opcao_exibicao"].iloc[0]
            st.rerun()
        else:
            st.warning("⚠️ Preencha todos os campos obrigatórios do insumo.")

def editar_insumo(index):
    """Carrega o insumo selecionado de volta nos campos para edição."""
    insumo = st.session_state.insumos[index]

    # Preenche novamente os campos
    st.session_state.descricao_exibicao = insumo["descricao"]
    st.session_state.codigo = insumo["codigo"]
    st.session_state.unidade = insumo["unidade"]
    st.session_state.quantidade = insumo["quantidade"]
    st.session_state.complemento = insumo["complemento"]

    # Marca o item sendo editado
    st.session_state.editando_insumo = index

# --- Renderiza tabela de insumos ---
if st.session_state.insumos:
    st.markdown("""
    <style>
    /* Cabeçalhos */
    .tabela-header {
        font-weight: 600;
        color: #333;
        border-bottom: 2px solid #ccc;
        padding-bottom: 2px;
        margin-bottom: 2px;
        font-size: 15px;
        display: flex;
        align-items: center;
        justify-content: flex-start;
    }
    .tabela-header.center {
        justify-content: center; /* centraliza só Qtd */
    }

    /* Linhas */
    .linha-insumo {
        border-bottom: 1px solid #e6e6e6;
        padding: 3px 0;
        font-size: 14px;
        line-height: 1.4;
        display: flex;
        align-items: center;
    }

    .center {
        justify-content: center;
        text-align: center;
    }

    /* Botão 🗑️ */
    div[data-testid="stButton"] button {
        border: none;
        background-color: transparent;
        color: #666;
        font-size: 18px;
        padding: 0;
        line-height: 1;
        transform: translateY(-2px);
    }

    div[data-testid="stButton"] button:hover {
        color: #d9534f;
        transform: scale(1.15) translateY(-2px);
    }
    </style>
    """, unsafe_allow_html=True)

    # Cabeçalho
    col1, col2, col3, col4 = st.columns([5.8, 1.2, 1.2, 0.5])
    with col1:
        st.markdown("<div class='tabela-header'>Insumos Adicionados</div>", unsafe_allow_html=True)
    with col2:
        st.markdown("<div class='tabela-header center'>Qtd</div>", unsafe_allow_html=True)
    with col3:
        st.markdown("<div class='tabela-header'>Unid</div>", unsafe_allow_html=True)
    with col4:
        st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)

    # Linhas
    for i, insumo in enumerate(st.session_state.insumos):
        col1, col2, col3, col4, col5 = st.columns([5.3, 1.2, 1.2, 0.5, 0.5])
        with col1:
            st.markdown(f"<div class='linha-insumo'>{insumo['descricao']}</div>", unsafe_allow_html=True)
        with col2:
            st.markdown(f"<div class='linha-insumo center'>{insumo['quantidade']}</div>", unsafe_allow_html=True)
        with col3:
            st.markdown(f"<div class='linha-insumo'>{insumo['unidade']}</div>", unsafe_allow_html=True)
        with col4:
            if st.button("✏️", key=f"edit_{i}"):
                editar_insumo(i)
                st.success(f"Insumo selecionado para edição: {insumo['descricao']}")
                st.rerun()
        with col5:
            if st.button("🗑️", key=f"delete_{i}"):
                st.session_state.insumos.pop(i)
                st.rerun()

# --- FINALIZAÇÃO DO PEDIDO ---
if st.button("📤 Enviar Pedido", use_container_width=True):
    campos_obrigatorios = [
        st.session_state.pedido_numero, st.session_state.data_pedido,
        st.session_state.solicitante, st.session_state.executivo,
        st.session_state.obra_selecionada, st.session_state.cnpj,
        st.session_state.endereco, st.session_state.cep, st.session_state.adm_obra
    ]
    if not all(campos_obrigatorios):
        st.warning("⚠️ Preencha todos os campos obrigatórios antes de enviar o pedido.")
        st.stop()
    if not st.session_state.insumos:
        st.warning("⚠️ Adicione pelo menos um insumo antes de enviar o pedido.")
        st.stop()

    erro = None
    ok = False
    with st.spinner("Enviando pedido e gerando arquivo... Aguarde!"):
        try:
            caminho_modelo = "Modelo_Pedido.xlsx"
            wb = load_workbook(caminho_modelo)
            ws = wb["Pedido"]

            ws["F2"] = st.session_state.pedido_numero
            ws["C3"] = st.session_state.data_pedido.strftime("%d/%m/%Y")
            ws["C4"] = st.session_state.solicitante
            ws["C5"] = st.session_state.executivo
            ws["C7"] = st.session_state.obra_selecionada
            ws["C8"] = st.session_state.cnpj
            ws["C9"] = st.session_state.endereco
            ws["C10"] = st.session_state.cep

            linha = 13
            for insumo in st.session_state.insumos:
                ws[f"B{linha}"] = insumo["codigo"]
                ws[f"C{linha}"] = insumo["descricao"]
                ws[f"D{linha}"] = insumo["unidade"]
                ws[f"E{linha}"] = insumo["quantidade"]
                ws[f"F{linha}"] = insumo["complemento"]
                linha += 1

            ultima_linha_util = linha - 1
            ws.print_area = f"A1:F{ultima_linha_util}"

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            st.session_state.excel_bytes = buffer.read()
            st.session_state.nome_arquivo = f"Pedido{st.session_state.pedido_numero} OC {st.session_state.obra_selecionada}.xlsx"

            enviar_email_pedido(
                f"Pedido{st.session_state.pedido_numero} OC {st.session_state.obra_selecionada}",
                st.session_state.excel_bytes,
                st.session_state.insumos,
                df_insumos
            )
            ok = True
        except Exception as e:
            erro = str(e)

    if ok:
        st.session_state.pedido_enviado = True
        st.success("✅ Pedido gerado e e-mail enviado com sucesso! Agora você pode baixar o arquivo Excel abaixo ⬇️")
    elif erro:
        st.error(f"❌ Erro ao gerar pedido: {erro}")

    # --- Salva o pedido no banco ---
    try:
        conn = sqlite3.connect("pedidos.db")
        c = conn.cursor()
        c.execute("""
            INSERT INTO pedidos (
                oc, pedido_numero, solicitante, executivo, obra,
                cnpj, endereco, cep, data_pedido, insumos, data_envio
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            st.session_state.get("login_oc"),
            st.session_state.pedido_numero,
            st.session_state.solicitante,
            st.session_state.executivo,
            st.session_state.obra_selecionada,
            st.session_state.cnpj,
            st.session_state.endereco,
            st.session_state.cep,
            st.session_state.data_pedido.strftime("%Y-%m-%d"),
            json.dumps(st.session_state.insumos, ensure_ascii=False),
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))
        conn.commit()
        conn.close()
        print("📦 Pedido salvo no banco de dados.")
    except Exception as e:
        print(f"Erro ao salvar pedido: {e}")

# --- BOTÕES APÓS ENVIO ---
if st.session_state.get("excel_bytes"):  # só renderiza se o arquivo existir
    col1, col2 = st.columns(2)

    with col1:
        if st.download_button(
            "📥 Baixar Excel",
            data=st.session_state.excel_bytes,
            file_name=st.session_state.nome_arquivo or "Pedido.xlsx",
            use_container_width=True
        ):
            # 🔹 Marca flags para limpar no próximo ciclo
            st.session_state.rerun_depois_download = True

    with col2:
        if st.button("🔄 Novo Pedido", use_container_width=True):
            for campo in [
                "pedido_numero", "solicitante", "executivo", "obra_selecionada",
                "cnpj", "endereco", "cep", "data_pedido",
                "excel_bytes", "nome_arquivo", "pedido_enviado"
            ]:
                if campo in st.session_state:
                    del st.session_state[campo]
            st.session_state.insumos = []
            st.rerun()

# --- 🔄 KEEP-ALIVE (mantém app ativo no Streamlit Cloud) ---
st.components.v1.html("""
<script>
setInterval(() => {
    fetch(window.location.pathname + '_stcore/health');
}, 120000);
</script>

""", height=0)
