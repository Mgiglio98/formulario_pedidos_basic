import streamlit as st
from datetime import date
from openpyxl import load_workbook
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import smtplib
from io import BytesIO

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Pedido de Materiais", page_icon="📦")

# --- CSS (espaçamento) ---
st.markdown("""
<style>
[data-testid="stAppViewContainer"] .main .block-container {
    padding-top: 0.5rem;
    padding-bottom: 2rem;
}
</style>
""", unsafe_allow_html=True)

# --- INICIALIZAÇÃO DE SESSÃO ---
if "insumos" not in st.session_state:
    st.session_state.insumos = []
if "resetar_insumo" not in st.session_state:
    st.session_state.resetar_insumo = False
if "resetar_pedido" not in st.session_state:
    st.session_state.resetar_pedido = False
if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None
if "nome_arquivo" not in st.session_state:
    st.session_state.nome_arquivo = ""
if "quantidade" not in st.session_state:
    st.session_state.quantidade = 1
if "descricao_exibicao" not in st.session_state:
    st.session_state.descricao_exibicao = ""
if "tipo_processo" not in st.session_state:
    st.session_state.tipo_processo = "Pedido de Materiais"
if "num_of_mae" not in st.session_state:
    st.session_state.num_of_mae = ""
if "fornecedor_of_filha" not in st.session_state:
    st.session_state.fornecedor_of_filha = ""


# --- RERUN APÓS DOWNLOAD ---
if st.session_state.get("rerun_depois_download", False):
    st.session_state.rerun_depois_download = False
    for campo in [
        "pedido_numero", "solicitante", "executivo", "obra_selecionada",
        "cnpj", "endereco", "cep", "data_pedido", "excel_bytes",
        "nome_arquivo", "pedido_enviado"
    ]:
        if campo in st.session_state:
            del st.session_state[campo]
    st.session_state.insumos = []
    st.rerun()

# --- CAMPOS PADRÃO ---
for campo in ["pedido_numero", "solicitante", "executivo", "obra_selecionada", "cnpj", "endereco", "cep"]:
    if campo not in st.session_state:
        st.session_state[campo] = ""

if "data_pedido" not in st.session_state:
    st.session_state.data_pedido = date.today()

# --- MAPA DE E-MAILS DOS ADMINISTRATIVOS ---
ADM_EMAILS = {
    "Maria Eduarda": "maria.eduarda@osborne.com.br",
    "Joice": "joice.oliveira@osborne.com.br",
    "Micaele": "micaele.ferreira@osborne.com.br",
    "Graziele": "graziele.horacio@osborne.com.br",
    "Roberto": "roberto.santos@osborne.com.br"
}

# --- FUNÇÕES AUXILIARES ---
def enviar_email_pedido(assunto, arquivo_bytes, insumos_adicionados, adm_emails, anexos=None):
    """Envia um único e-mail do pedido, com cópia fixa e variável, e aviso se houver insumos sem código."""
    smtp_server = "smtp.office365.com"
    smtp_port = 587
    smtp_user = "matheus.almeida@osborne.com.br"
    smtp_password = st.secrets["SMTP_PASSWORD"]

    # --- Endereços de cópia ---
    cc_addr = ["joice.oliveira@osborne.com.br"]  # cópia fixa
    adm_email = adm_emails.get(st.session_state.get("adm_obra"))
    if adm_email and adm_email not in cc_addr:
        cc_addr.append(adm_email)

    # --- Identifica insumos sem código ---
    sem_codigo = [
        f"{item['descricao']} — {item['quantidade']}"
        for item in insumos_adicionados
        if not item.get("codigo") or str(item["codigo"]).strip() == ""
    ]

    # --- Dados básicos do formulário para resumo ---
    tipo_proc = st.session_state.get("tipo_processo", "Pedido de Materiais")
    pedido_num = st.session_state.get("pedido_numero", "")
    obra = st.session_state.get("obra_selecionada", "")
    solicitante = st.session_state.get("solicitante", "")
    executivo = st.session_state.get("executivo", "")
    data_pedido = st.session_state.get("data_pedido", date.today())

    try:
        data_fmt = data_pedido.strftime("%d/%m/%Y")
    except Exception:
        data_fmt = str(data_pedido)

    sem_codigo_texto = ""
    if sem_codigo:
        sem_codigo_texto = "\n\nInsumos sem código cadastrado:\n" + "\n".join(f"- {linha}" for linha in sem_codigo)

    # --- Corpo do e-mail por tipo de processo ---
    if tipo_proc == "Pedido de Materiais":
        corpo_email = f"""Olá! Novo PEDIDO DE MATERIAIS recebido ✅

Resumo do pedido:
- Nº Pedido: {pedido_num}
- Obra: {obra}
- Solicitante: {solicitante}
- Executivo: {executivo}
- Data: {data_fmt}

Esse pedido deve ser conferido e, se estiver de acordo, seguirá para Requisição e OF.{sem_codigo_texto}
"""

    elif tipo_proc == "Requisição para Cotação":
        corpo_email = f"""Olá! Nova COTAÇÃO recebida ✅

Resumo da solicitação:
- Referência: {pedido_num}
- Obra: {obra}
- Solicitante: {solicitante}
- Executivo: {executivo}
- Data: {data_fmt}

As propostas/orçamentos enviados pela obra estão anexos a este e-mail.
Utilizar este pedido como base para análise das cotações e definição do fornecedor.{sem_codigo_texto}
"""

    elif tipo_proc == "Criação de ED":
        num_of_mae = st.session_state.get("num_of_mae", "")
        fornecedor_of_filha = st.session_state.get("fornecedor_of_filha", "")

        corpo_email = f"""Olá! Nova SOLICITAÇÃO DE ED recebida ✅

Resumo da solicitação:
- Referência: {pedido_num}
- Obra: {obra}
- Solicitante: {solicitante}
- Executivo: {executivo}
- Data: {data_fmt}
- Nº OF Mãe: {num_of_mae}
- Fornecedor da OF filha: {fornecedor_of_filha}

Favor analisar e, se estiver de acordo, proceder com a criação da Requisição da ED e OF filha no sistema, vinculando à OF Mãe informada.{sem_codigo_texto}
"""
    else:
        # fallback (não deveria acontecer, mas deixa robusto)
        corpo_email = f"""Olá! Novo formulário recebido ✅

Tipo de processo selecionado: {tipo_proc or "Não informado"}{sem_codigo_texto}
"""

    # --- Montagem do e-mail ---
    msg = MIMEMultipart()
    msg["From"] = smtp_user
    msg["To"] = smtp_user
    msg["Cc"] = ", ".join(cc_addr)
    msg["Subject"] = assunto
    msg.attach(MIMEText(corpo_email, "plain"))

    # --- Anexo principal (planilha do formulário) ---
    anexo = MIMEApplication(
        arquivo_bytes,
        _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    nome_arquivo = st.session_state.nome_arquivo or "Pedido.xlsx"
    anexo.add_header('Content-Disposition', 'attachment', filename=nome_arquivo)
    msg.attach(anexo)

    # --- Anexos extras (cotação / ED) ---
    if anexos:
        for arquivo in anexos:
            try:
                conteudo = arquivo.getvalue()
                anexo_extra = MIMEApplication(conteudo)
                anexo_extra.add_header(
                    'Content-Disposition',
                    'attachment',
                    filename=arquivo.name
                )
                msg.attach(anexo_extra)
            except Exception as e:
                print(f"Erro ao anexar arquivo extra {getattr(arquivo, 'name', 'sem_nome')}: {e}")

    # --- Envio ---
    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
            print("📨 E-mail enviado com sucesso!")
    except Exception as e:
        print(f"Erro ao enviar e-mail: {e}")

def carregar_pedido_existente(arquivo):
    """Lê um Excel de pedido gerado pelo formulário e recarrega os campos no session_state."""
    try:
        wb = load_workbook(arquivo)
        ws = wb["Pedido"]
    except Exception as e:
        st.error(f"Não foi possível ler o arquivo de pedido: {e}")
        return

    # --- Cabeçalho (mesmos campos que você grava ao gerar) ---
    st.session_state.pedido_numero = str(ws["F2"].value or "").strip()

    data_raw = ws["C3"].value
    data_pedido = pd.to_datetime(data_raw, errors="coerce")
    if pd.isna(data_pedido):
        data_pedido = date.today()
    else:
        data_pedido = data_pedido.date()
    st.session_state.data_pedido = data_pedido

    st.session_state.solicitante      = str(ws["C4"].value or "").strip()
    st.session_state.executivo        = str(ws["C5"].value or "").strip()
    st.session_state.obra_selecionada = str(ws["C7"].value or "").strip()
    st.session_state.cnpj             = str(ws["C8"].value or "").strip()
    st.session_state.endereco         = str(ws["C9"].value or "").strip()
    st.session_state.cep              = str(ws["C10"].value or "").strip()

    # Campos que não vêm do Excel
    st.session_state.adm_obra = ""
    st.session_state.num_of_mae = ""
    st.session_state.fornecedor_of_filha = ""

    # --- Itens do pedido (a partir da linha 13) ---
    st.session_state.insumos = []
    linha = 13
    while True:
        cod   = ws[f"B{linha}"].value
        desc  = ws[f"C{linha}"].value
        unid  = ws[f"D{linha}"].value
        qtd   = ws[f"E{linha}"].value
        compl = ws[f"F{linha}"].value

        # linha totalmente vazia -> fim
        if not any([cod, desc, unid, qtd, compl]):
            break

        if desc and qtd:
            try:
                qtd_int = int(qtd)
            except Exception:
                try:
                    qtd_int = int(float(qtd))
                except Exception:
                    qtd_int = qtd

            novo_insumo = {
                "descricao": str(desc).strip(),
                "codigo": "" if cod is None else str(cod).strip(),
                "unidade": "" if unid is None else str(unid).strip(),
                "quantidade": qtd_int,
                "complemento": "" if compl is None else str(compl).strip(),
            }
            st.session_state.insumos.append(novo_insumo)

        linha += 1

def carregar_dados():
    """Carrega dados de empreendimentos e insumos."""
    df_empreend = pd.read_excel("Empreendimentos.xlsx")
    df_empreend.columns = df_empreend.columns.str.strip().str.upper()

    df_insumos = pd.read_excel("Insumos.xlsx")
    df_insumos["Min"] = pd.to_numeric(df_insumos.iloc[:, 3], errors="coerce")
    df_insumos["Max"] = pd.to_numeric(df_insumos.iloc[:, 4], errors="coerce")
    df_insumos["Basico"] = df_insumos["Min"].notna() & df_insumos["Max"].notna()
    df_insumos = df_insumos[df_insumos["Descrição"].notna() & (df_insumos["Descrição"].str.strip() != "")]

    df_empreend.loc[-1] = [""] * df_empreend.shape[1]
    df_empreend.index = df_empreend.index + 1
    df_empreend = df_empreend.sort_index()

    insumos_vazios = pd.DataFrame({"Código": [""], "Descrição": [""], "Unidade": [""]})
    df_insumos = pd.concat([insumos_vazios, df_insumos], ignore_index=True)
    df_insumos["Código"] = df_insumos["Código"].fillna("").astype(str)
    return df_empreend, df_insumos

# --- CARREGAMENTO DE DADOS ---
df_empreend, df_insumos = carregar_dados()

# --- LIMPEZA APÓS ENVIO ---
if st.session_state.get("limpar_pedido", False):
    for campo in ["pedido_numero", "solicitante", "executivo", "obra_selecionada", "cnpj", "endereco", "cep"]:
        if campo in st.session_state:
            try:
                st.session_state[campo] = ""
            except Exception:
                pass
    st.session_state.data_pedido = date.today()
    st.session_state.insumos = []
    st.session_state.limpar_pedido = False

# --- LOGO E CABEÇALHO ---
col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    st.image("logo.png", width=300)

st.markdown("""
<div style='text-align: center;'>
    <p style='font-size: 14px; color: #555;'>
        Preencha os campos com atenção. Verifique se todos os dados estão corretos antes de enviar.<br>
        Ao finalizar, a solicitação será automaticamente enviado para o e-mail do setor de Suprimentos.<br>
        Você poderá baixar a planilha gerada após o envio, para registro ou controle.
    <h2 style='color: #000000;'>Suprimentos - Osborne</h2>
    </p>
</div>
""", unsafe_allow_html=True)

# --- TIPO DE PROCESSO (PEDIDO / COTAÇÃO / ED) ---
# --- TIPO DE PROCESSO (PEDIDO / COTAÇÃO / ED / CORRIGIR) ---
#st.markdown("### Tipo de processo")

TIPO_PEDIDO   = "Pedido de Materiais"
TIPO_COTACAO  = "Requisição para Cotação"
TIPO_ED       = "Criação de ED"
TIPO_CORRIGIR = "Corrigir Pedido"

opcoes_tipo = [TIPO_PEDIDO, TIPO_COTACAO, TIPO_ED, TIPO_CORRIGIR]

# valor atual na sessão (se não tiver, usa Pedido como padrão)
valor_atual = st.session_state.get("tipo_processo", TIPO_PEDIDO)
if valor_atual not in opcoes_tipo:
    valor_atual = TIPO_PEDIDO

tipo_processo = st.radio(
    "Selecione o processo para este formulário:",
    options=opcoes_tipo,
    index=opcoes_tipo.index(valor_atual),
    horizontal=True
)

st.session_state["tipo_processo"] = tipo_processo

st.divider()

# --- CORRIGIR PEDIDO (upload vem antes de Dados do Pedido) ---
if st.session_state.tipo_processo == TIPO_CORRIGIR:
    st.markdown("#### Corrigir pedido existente")
    st.write("Envie abaixo o arquivo Excel de um pedido gerado anteriormente pelo formulário para carregar os dados novamente.")

    arquivo_corrigir = st.file_uploader(
        "Arquivo do pedido (Excel gerado pelo formulário)",
        type=["xlsx"],
        key="arquivo_corrigir"
    )

    if arquivo_corrigir is not None:
        if st.button("Carregar pedido para edição"):
            carregar_pedido_existente(arquivo_corrigir)
            # depois de carregar, voltamos o tipo para Pedido de Materiais
            st.session_state["tipo_processo"] = TIPO_PEDIDO
            st.success("✅ Pedido carregado! Revise os dados, ajuste o que for necessário e clique em 'Enviar Pedido'.")
            st.rerun()

    st.markdown("---")

# --- DADOS DO PEDIDO ---
with st.expander("📋 Dados do Pedido", expanded=True):
    if st.session_state.resetar_pedido:
        st.session_state.pedido_numero = ""
        st.session_state.data_pedido = date.today()
        st.session_state.solicitante = ""
        st.session_state.executivo = ""
        st.session_state.adm_obra = ""
        st.session_state.obra_selecionada = ""
        st.session_state.cnpj = ""
        st.session_state.endereco = ""
        st.session_state.cep = ""
        st.session_state.resetar_pedido = False

    col1, col2 = st.columns(2)
    with col1:
        pedido_numero = st.text_input("Pedido Nº", key="pedido_numero")
        solicitante = st.text_input("Solicitante", key="solicitante")
        obra_selecionada = st.selectbox("Obra", df_empreend["EMPREENDIMENTO"].unique(), index=0, key="obra_selecionada")
    with col2:
        data_pedido = st.date_input(
            "Data",
            key="data_pedido",
            value=st.session_state.data_pedido if "data_pedido" in st.session_state else date.today()
        )
        executivo = st.text_input("Executivo", key="executivo")
        
        opcoes_adm = [""] + list(ADM_EMAILS.keys())  # primeira opção em branco
        adm_obra = st.selectbox(
            "Administrativo da Obra",
            opcoes_adm,
            index=0,
            key="adm_obra"
        )

    if obra_selecionada:
        filtro = df_empreend["EMPREENDIMENTO"] == obra_selecionada
        if filtro.any():
            dados_obra = df_empreend.loc[filtro].iloc[0]
            st.session_state.cnpj = dados_obra["CNPJ"]
            st.session_state.endereco = dados_obra["ENDERECO"]
            st.session_state.cep = dados_obra["CEP"]
    
    st.text_input("CNPJ/CPF", key="cnpj", disabled=True)
    st.text_input("Endereço", key="endereco", disabled=True)
    st.text_input("CEP", value=st.session_state.get("cep", ""), disabled=True)

# --- CAMPOS ESPECÍFICOS POR TIPO DE PROCESSO ---
anexos_processo = []  # default, para usar mais à frente

# garante que os campos extras existam na sessão
if "num_of_mae" not in st.session_state:
    st.session_state.num_of_mae = ""
if "fornecedor_of_filha" not in st.session_state:
    st.session_state.fornecedor_of_filha = ""

if st.session_state.tipo_processo == TIPO_COTACAO:
    # --------- COTAÇÃO ---------
    with st.expander("📎 Propostas / Orçamentos (Cotação)", expanded=True):
        st.write("Anexe aqui as propostas recebidas para esta cotação.")
        anexos_processo = st.file_uploader(
            "Selecionar arquivos de proposta",
            type=["pdf", "xlsx", "xls", "csv", "png", "jpg", "jpeg"],
            accept_multiple_files=True,
            key="anexos_cotacao"
        )
        if anexos_processo:
            st.info(f"{len(anexos_processo)} arquivo(s) será(ão) enviado(s) junto com a requisição de cotação.")

elif st.session_state.tipo_processo == TIPO_ED:
    # --------- ED / OF FILHA ---------
    with st.expander("📄 Dados da ED / OF filha", expanded=True):
        st.text_input("Nº OF Mãe", key="num_of_mae")
        st.text_input("Fornecedor da OF filha", key="fornecedor_of_filha")

    with st.expander("📎 Documentos da ED / OF filha", expanded=False):
        anexos_processo = st.file_uploader(
            "Anexar documentos (planilhas, PDFs, prints, etc.)",
            type=["pdf", "xlsx", "xls", "csv", "png", "jpg", "jpeg"],
            accept_multiple_files=True,
            key="anexos_ed"
        )
        if anexos_processo:
            st.info(f"{len(anexos_processo)} arquivo(s) será(ão) enviado(s) junto com a requisição de ED.")

# --- ADIÇÃO DE INSUMOS ---
with st.expander("➕ Adicionar Insumo", expanded=True):

    if st.session_state.get("limpar_campos_insumo", False):
        # 🔹 Remove todos os valores dos campos
        for campo in ["descricao_exibicao", "descricao_livre", "codigo", "unidade", "quantidade", "complemento"]:
            if campo in st.session_state:
                del st.session_state[campo]

        # 🔹 Garante valor padrão inicial
        st.session_state.quantidade = 1
        st.session_state.descricao_exibicao = ""
        st.session_state.complemento = ""
        st.session_state.limpar_campos_insumo = False
        st.rerun()  # 🔁 força recarregar já limpo
    
    df_insumos_lista = df_insumos.sort_values(by="Descrição", ascending=True).copy()
    df_insumos_lista["opcao_exibicao"] = df_insumos_lista.apply(
        lambda x: f"{x['Descrição']} – {x['Código']} ({x['Unidade']})" if pd.notna(x["Código"]) and str(x["Código"]).strip() != "" else x["Descrição"],
        axis=1
    )
    
    descricao_exibicao = st.selectbox(
        "Descrição do insumo (Digite em MAIÚSCULO)",
        df_insumos_lista["opcao_exibicao"],
        key="descricao_exibicao"
    )

    dados_insumo = df_insumos_lista[df_insumos_lista["opcao_exibicao"] == descricao_exibicao].iloc[0]
    codigo_sel = str(dados_insumo["Código"]).strip()
    usando_base = codigo_sel != ""

    if usando_base:
        st.session_state.codigo = dados_insumo["Código"]
        st.session_state.unidade = dados_insumo["Unidade"]
        st.session_state.descricao = dados_insumo["Descrição"]
    else:
        st.session_state.codigo = ""
        st.session_state.descricao = ""

    if "unidade" not in st.session_state or not st.session_state.unidade:
        st.session_state.unidade = ""

    st.write("Ou preencha manualmente o Nome e Unidade se não estiver listado:")

    descricao_livre = st.text_input("Nome do insumo (livre)", key="descricao_livre", disabled=usando_base)
    st.text_input("Código do insumo", key="codigo", disabled=True)
    st.text_input("Unidade", key="unidade", disabled=usando_base)
    quantidade = st.number_input("Quantidade", min_value=1, step=1, format="%d", key="quantidade")
    complemento = st.text_area(
        "Complemento, se necessário (Utilize para especificar medidas, marcas, cores e/ou tamanhos)",
        key="complemento"
    )

    if st.button("➕ Adicionar insumo"):
        descricao_final = st.session_state.descricao if usando_base else descricao_livre
    
        if descricao_final and quantidade > 0 and (usando_base or st.session_state.unidade.strip()):
            novo_insumo = {
                "descricao": descricao_final,
                "codigo": st.session_state.codigo if usando_base else "",
                "unidade": st.session_state.unidade,
                "quantidade": quantidade,
                "complemento": complemento,
            }
            st.session_state.insumos.append(novo_insumo)
    
            # 🔹 Marca para limpar na próxima renderização
            st.session_state.limpar_campos_insumo = True
    
            st.success("✅ Insumo adicionado com sucesso!")
            st.rerun()
        
        else:
            st.warning("⚠️ Preencha todos os campos obrigatórios do insumo.")

# --- Renderiza tabela de insumos ---
if st.session_state.insumos:
    st.markdown("""
    <style>
    /* Cabeçalhos */
    .tabela-header {
        font-weight: 600;
        color: #333;
        border-bottom: 2px solid #ccc;
        padding-bottom: 2px;
        margin-bottom: 2px;
        font-size: 15px;
        display: flex;
        align-items: center;
        justify-content: flex-start;
    }
    .tabela-header.center {
        justify-content: center; /* centraliza só Qtd */
    }

    /* Linhas */
    .linha-insumo {
        border-bottom: 1px solid #e6e6e6;
        padding: 3px 0;
        font-size: 14px;
        line-height: 1.4;
        display: flex;
        align-items: center;
    }

    .center {
        justify-content: center;
        text-align: center;
    }

    /* Botão 🗑️ */
    div[data-testid="stButton"] button {
        border: none;
        background-color: transparent;
        color: #666;
        font-size: 18px;
        padding: 0;
        line-height: 1;
        transform: translateY(-2px);
    }

    div[data-testid="stButton"] button:hover {
        color: #d9534f;
        transform: scale(1.15) translateY(-2px);
    }
    </style>
    """, unsafe_allow_html=True)

    # Cabeçalho
    col1, col2, col3, col4 = st.columns([5.8, 1.2, 1.2, 0.5])
    with col1:
        st.markdown("<div class='tabela-header'>Insumos Adicionados</div>", unsafe_allow_html=True)
    with col2:
        st.markdown("<div class='tabela-header center'>Qtd</div>", unsafe_allow_html=True)
    with col3:
        st.markdown("<div class='tabela-header'>Unid</div>", unsafe_allow_html=True)
    with col4:
        st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)

    # Linhas
    for i, insumo in enumerate(st.session_state.insumos):
        col1, col2, col3, col4 = st.columns([5.8, 1.2, 1.2, 0.5])
        with col1:
            st.markdown(f"<div class='linha-insumo'>{insumo['descricao']}</div>", unsafe_allow_html=True)
        with col2:
            st.markdown(f"<div class='linha-insumo center'>{insumo['quantidade']}</div>", unsafe_allow_html=True)
        with col3:
            st.markdown(f"<div class='linha-insumo'>{insumo['unidade']}</div>", unsafe_allow_html=True)
        with col4:
            if st.button("🗑️", key=f"delete_{i}"):
                st.session_state.insumos.pop(i)
                st.rerun()

# --- FINALIZAÇÃO DO PEDIDO ---
if st.button("📤 Enviar Pedido", use_container_width=True):

    tipo_proc = st.session_state.get("tipo_processo", TIPO_PEDIDO)
    
    campos_obrigatorios = [
        st.session_state.pedido_numero, st.session_state.data_pedido,
        st.session_state.solicitante, st.session_state.executivo,
        st.session_state.obra_selecionada, st.session_state.cnpj,
        st.session_state.endereco, st.session_state.cep, st.session_state.adm_obra
    ]
    if not st.session_state.adm_obra or st.session_state.adm_obra.strip() == "":
        st.warning("⚠️ Selecione o Administrativo da obra antes de enviar o pedido.")
        st.stop()
    if not all(campos_obrigatorios):
        st.warning("⚠️ Preencha todos os campos obrigatórios antes de enviar o pedido.")
        st.stop()
    if not st.session_state.insumos:
        st.warning("⚠️ Adicione pelo menos um insumo antes de enviar o pedido.")
        st.stop()

    if tipo_proc == TIPO_COTACAO:
        if not anexos_processo:
            st.warning("⚠️ Para 'Requisição para Cotação', anexe pelo menos uma proposta/orçamento.")
            st.stop()

    if tipo_proc == TIPO_ED:
        num_of_mae = st.session_state.get("num_of_mae", "").strip()
        fornecedor_of_filha = st.session_state.get("fornecedor_of_filha", "").strip()

        if not num_of_mae:
            st.warning("⚠️ Informe o Nº da OF Mãe para criar a ED / OF filha.")
            st.stop()

        if not fornecedor_of_filha:
            st.warning("⚠️ Informe o Fornecedor da OF filha.")
            st.stop()

        if not anexos_processo:
            st.warning("⚠️ Anexe pelo menos um documento para a ED / OF filha.")
            st.stop()
    
    erro = None
    ok = False
    with st.spinner("Enviando pedido e gerando arquivo... Aguarde!"):
        try:
            caminho_modelo = "Modelo_Pedido.xlsx"
            wb = load_workbook(caminho_modelo)
            ws = wb["Pedido"]

            ws["F2"] = st.session_state.pedido_numero
            ws["C3"] = st.session_state.data_pedido.strftime("%d/%m/%Y")
            ws["C4"] = st.session_state.solicitante
            ws["C5"] = st.session_state.executivo
            ws["C7"] = st.session_state.obra_selecionada
            ws["C8"] = st.session_state.cnpj
            ws["C9"] = st.session_state.endereco
            ws["C10"] = st.session_state.cep

            linha = 13
            for insumo in st.session_state.insumos:
                ws[f"B{linha}"] = insumo["codigo"]
                ws[f"C{linha}"] = insumo["descricao"]
                ws[f"D{linha}"] = insumo["unidade"]
                ws[f"E{linha}"] = insumo["quantidade"]
                ws[f"F{linha}"] = insumo["complemento"]
                linha += 1

            ultima_linha_util = linha - 1
            ws.print_area = f"A1:F{ultima_linha_util}"

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            st.session_state.excel_bytes = buffer.read()
            st.session_state.nome_arquivo = f"Pedido{st.session_state.pedido_numero} OC {st.session_state.obra_selecionada}.xlsx"

            enviar_email_pedido(
                f"Pedido{st.session_state.pedido_numero} OC {st.session_state.obra_selecionada}",
                st.session_state.excel_bytes,
                st.session_state.insumos,
                ADM_EMAILS,
                anexos=anexos_processo
            )
            ok = True
        except Exception as e:
            erro = str(e)

    if ok:
        st.session_state.pedido_enviado = True
        st.success("✅ Pedido gerado e e-mail enviado com sucesso! Agora você pode baixar o arquivo Excel abaixo ⬇️")
    elif erro:
        st.error(f"❌ Erro ao gerar pedido: {erro}")

# --- BOTÕES APÓS ENVIO ---
if st.session_state.get("excel_bytes"):  # só renderiza se o arquivo existir
    col1, col2 = st.columns(2)

    # Botão de download do Excel
    with col1:
        if st.download_button(
            "📥 Baixar Excel",
            data=st.session_state.excel_bytes,
            file_name=st.session_state.nome_arquivo or "Pedido.xlsx",
            use_container_width=True
        ):
            # 🔹 Marca flags para limpar no próximo ciclo
            st.session_state.rerun_depois_download = True

    # Botão de novo pedido
    with col2:
        if st.button("🔄 Novo Pedido", use_container_width=True):
            for campo in [
                "pedido_numero", "solicitante", "executivo", "obra_selecionada",
                "cnpj", "endereco", "cep", "data_pedido",
                "excel_bytes", "nome_arquivo", "pedido_enviado",
                "adm_obra",              # 👈 limpa o ADM
                "num_of_mae",
                "fornecedor_of_filha",
                "tipo_processo",
                "anexos_cotacao",
                "anexos_ed",
            ]:
                if campo in st.session_state:
                    del st.session_state[campo]

            # garante defaults para próximo carregamento
            st.session_state.insumos = []
            st.session_state["adm_obra"] = ""        # 👈 volta select p/ branco
            st.session_state["tipo_processo"] = TIPO_PEDIDO

            st.rerun()

# --- 🔄 KEEP-ALIVE (mantém app ativo no Streamlit Cloud) ---
st.components.v1.html("""
<script>
setInterval(() => {
    fetch(window.location.pathname + '_stcore/health');
}, 120000);
</script>
""", height=0)
