import streamlit as st
from datetime import date
from openpyxl import load_workbook
import pandas as pd
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import smtplib
from io import BytesIO
from datetime import datetime

# --- CONFIGURAÇÃO DA PÁGINA ---
st.set_page_config(page_title="Pedido de Materiais", page_icon="📦")

# --- CSS (espaçamento) ---
st.markdown("""
<style>
[data-testid="stAppViewContainer"] .main .block-container {
    padding-top: 0.5rem;
    padding-bottom: 2rem;
}
</style>
""", unsafe_allow_html=True)

# --- INICIALIZAÇÃO DE SESSÃO ---
if "insumos" not in st.session_state:
    st.session_state.insumos = []
if "resetar_insumo" not in st.session_state:
    st.session_state.resetar_insumo = False
if "resetar_pedido" not in st.session_state:
    st.session_state.resetar_pedido = False
if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None
if "nome_arquivo" not in st.session_state:
    st.session_state.nome_arquivo = ""
if "quantidade" not in st.session_state:
    st.session_state.quantidade = 1.0
if "descricao_exibicao" not in st.session_state:
    st.session_state.descricao_exibicao = ""
if "tipo_processo" not in st.session_state:
    st.session_state.tipo_processo = "Pedido de Materiais"
if "num_of_mae" not in st.session_state:
    st.session_state.num_of_mae = ""
if "fornecedor_of_filha" not in st.session_state:
    st.session_state.fornecedor_of_filha = ""

# --- RERUN APÓS DOWNLOAD ---
if st.session_state.get("rerun_depois_download", False):
    st.session_state.rerun_depois_download = False
    for campo in [
        "pedido_numero", "solicitante", "executivo", "obra_selecionada",
        "cnpj", "endereco", "cep", "data_pedido", "excel_bytes",
        "nome_arquivo", "pedido_enviado",
        "adm_obra", "num_of_mae", "fornecedor_of_filha",
        "tipo_processo", "anexos_ed"
    ]:
        if campo in st.session_state:
            del st.session_state[campo]
    st.session_state.insumos = []
    st.session_state["tipo_processo"] = "Pedido de Materiais"
    st.rerun()

def validar_data_br(txt: str):
    txt = (txt or "").strip()
    if not txt:
        return None, "Preencha a data de necessidade do insumo (DD/MM/YYYY)."
    try:
        dt = datetime.strptime(txt, "%d/%m/%Y").date()
        return dt, None
    except ValueError:
        return None, "Data inválida. Use o formato DD/MM/YYYY."

# --- CAMPOS PADRÃO ---
for campo in ["pedido_numero", "solicitante", "executivo", "obra_selecionada", "cnpj", "endereco", "cep"]:
    if campo not in st.session_state:
        st.session_state[campo] = ""

if "data_pedido" not in st.session_state:
    st.session_state.data_pedido = date.today()

# --- MAPA DE E-MAILS DOS ADMINISTRATIVOS ---
ADM_EMAILS = {
    "Maria Eduarda": "maria.eduarda@osborne.com.br",
    "Joice": "joice.oliveira@osborne.com.br",
    "Micaele": "micaele.ferreira@osborne.com.br",
    "Graziele": "graziele.horacio@osborne.com.br",
    "Roberto": "roberto.santos@osborne.com.br"
}

OBRA_EXECUTIVOS = {
    "2317 - LUIZ ALBERTO HESS BORGES": [
        {"executivo": "Julia Vigorito", "email": "julia.vigorito@osborne.com.br"},
    ],
    "2514 - FELIPE HESS BORGES": [
        {"executivo": "Matheus Sanches", "email": "matheus.sanches@osborne.com.br"},
    ],
    "2411 - JOÃO CARLOS BEHISNELIAN": [
        {"executivo": "Vitor Ramos", "email": "vitor.ramos@osborne.com.br"},
    ],
    "2407 - SUN MORITZ ADMINISTRADORA": [
        {"executivo": "Carolina Oliveira", "email": "carolina.oliveira@osborne.com.br"},
    ],
    "2511 - 1807 PARTICIPAÇÕES LTDA": [
        {"executivo": "Andre Pestana", "email": "andre.pestana@osborne.com.br"},
    ],
    "2512 - ROBERTO KLABIN MARTINS XAVIER": [
        {"executivo": "Leonardo Devico", "email": "leonardo.devico@osborne.com.br"},
    ],
    "2516 - JOSÉ CARLOS MORAES ABREU FILHO": [
        {"executivo": "Andre Pestana", "email": "andre.pestana@osborne.com.br"},
    ],
    "2505 - EW ADMINISTRADORA LTDA": [
        {"executivo": "Felipe Duarte", "email": "felipe.duarte@osborne.com.br"},
    ],
    "2316 - MARCO AURÉLIO SIMÃO FREIRE": [
        {"executivo": "Caio Fausto", "email": "caio.fausto@osborne.com.br"},
    ],
    "2504 - MARIA ANGÉLICA A. M. DA COSTA": [
        {"executivo": "Alberto Teixeira", "email": "alberto.teixeira@osborne.com.br"},
    ],
    "2509 - RAFAEL CURSINO DE MOURA LEVY": [
        {"executivo": "Caio Fausto", "email": "caio.fausto@osborne.com.br"},
    ],
    "2510 - SAMAUMA EVENTOS LTDA": [
        {"executivo": "Alberto Teixeira", "email": "alberto.teixeira@osborne.com.br"},
    ],
    "2515 - MARCO FREIRE (ÁREA EXTERNA)": [
        {"executivo": "Caio Fausto", "email": "caio.fausto@osborne.com.br"},
    ],
    "2506 - KATIA FERREIRA DE BARROS": [
        {"executivo": "Vitor Carvalho", "email": "vitor.carvalho@osborne.com.br"},
        {"executivo": "Danielle Monteiro", "email": "danielle.monteiro@osborne.com.br"},
    ],
    "2507 - KATIA FERREIRA DE BARROS": [
        {"executivo": "Vitor Carvalho", "email": "vitor.carvalho@osborne.com.br"},
        {"executivo": "Danielle Monteiro", "email": "danielle.monteiro@osborne.com.br"},
    ],
    "2212 - IDEA INVEST. IMOBILIÁRIOS LTDA.": [
        {"executivo": "Giullian Moura", "email": "giullian.moura@osborne.com.br"},
        {"executivo": "Luciana Abreu", "email": "luciana.abreu@osborne.com.br"},
    ],
    "2409 - MARIA BELTRÃO SALDANHA COELHO": [
        {"executivo": "Najara Camargo", "email": "najara.camargo@osborne.com.br"},
    ],
    "2601 - REGINA CAMPOS SALLES MORAES ABREU": [
        {"executivo": "Marcos Varjão", "email": "marcos.varjao@osborne.com.br"},
    ],
    "2404 - LUIZ HENRIQUE FRAGA": [
        {"executivo": "Leno Fagundes", "email": "leno.fagundes@osborne.com.br"},
        {"executivo": "Sofia Goes", "email": "sofia.goes@osborne.com.br"},
    ],
    "2303 - LUIZ ROGÉRIO BERTO": [
        {"executivo": "Igor Bueno", "email": "igor.bueno@osborne.com.br"},
    ],
    "2603 - FERNANDO AUGUSTO COELHO F DE VASCONCELLOS": [
        {"executivo": "Najara Camargo", "email": "najara.camargo@osborne.com.br"},
    ],
    "2213 - TOMAS DA VEIGA PEREIRA": [
        {"executivo": "Najara Camargo", "email": "najara.camargo@osborne.com.br"},
    ],
}

# Lista única de executivos a partir do OBRA_EXECUTIVOS
EXECUTIVOS_OPCOES = sorted({
    item["executivo"].strip()
    for lista in OBRA_EXECUTIVOS.values()
    for item in lista
    if item.get("executivo") and str(item["executivo"]).strip()
})

OBRAS_SEM_EXECUTIVO_FIXO = {
    "9992 - GARANTIA DE OBRAS",
    "9991 - DÉBITO ADMINISTRAÇÃO (OBRAS)",
}

# --- FUNÇÕES AUXILIARES ---
def enviar_email_pedido(assunto, arquivo_bytes, insumos_adicionados, adm_emails, anexos=None):
    """Envia um único e-mail do pedido, com cópia fixa e variável, e aviso se houver insumos sem código."""
    smtp_server = "smtp.office365.com"
    smtp_port = 587
    smtp_user = "matheus.almeida@osborne.com.br"
    smtp_password = st.secrets["SMTP_PASSWORD"]

    cc_addr = ["antonio.macedo@osborne.com.br"]

    adm_email = adm_emails.get(st.session_state.get("adm_obra"))
    if adm_email and adm_email not in cc_addr:
        cc_addr.append(adm_email)

    exec_emails = st.session_state.get("exec_emails_obra", [])
    for e in exec_emails:
        e = (e or "").strip()
        if e and e not in cc_addr:
            cc_addr.append(e)

    sem_codigo = [
        f"{item['descricao']} — {item['quantidade']}"
        for item in insumos_adicionados
        if not item.get("codigo") or str(item["codigo"]).strip() == ""
    ]

    tipo_proc = st.session_state.get("tipo_processo", "Pedido de Materiais")
    pedido_num = st.session_state.get("pedido_numero", "")
    obra = st.session_state.get("obra_selecionada", "")
    solicitante = st.session_state.get("solicitante", "")
    executivo = st.session_state.get("executivo", "")
    data_pedido = st.session_state.get("data_pedido", date.today())

    try:
        data_fmt = data_pedido.strftime("%d/%m/%Y")
    except Exception:
        data_fmt = str(data_pedido)

    sem_codigo_texto = ""
    if sem_codigo:
        sem_codigo_texto = "\n\nInsumos sem código cadastrado:\n" + "\n".join(f"- {linha}" for linha in sem_codigo)

    if tipo_proc == "Pedido de Materiais":
        if sem_codigo:
            lista_formatada = "".join(f"- {item}\n" for item in sem_codigo)
            corpo_email = f"""
Olá! Novo pedido recebido ✅

Favor validar antes de criarmos a requisição.

Os seguintes insumos estão no pedido sem o código cadastrado:
{lista_formatada}
            """
        else:
            corpo_email = """
Olá! Novo pedido recebido ✅

Favor validar antes de criarmos a requisição.
            """

    elif tipo_proc == "Criação de ED":
        num_of_mae = st.session_state.get("num_of_mae", "")
        fornecedor_of_filha = st.session_state.get("fornecedor_of_filha", "")

        corpo_email = f"""Olá! Nova SOLICITAÇÃO DE ED recebida ✅

Resumo da solicitação:
- Referência: {pedido_num}
- Obra: {obra}
- Solicitante: {solicitante}
- Executivo: {executivo}
- Data: {data_fmt}
- Nº OF Mãe: {num_of_mae}
- Fornecedor da OF filha: {fornecedor_of_filha}

Favor analisar e, se estiver de acordo, proceder com a criação da Requisição da ED e OF filha no sistema, vinculando à OF Mãe informada.{sem_codigo_texto}
"""
    else:
        corpo_email = f"""Olá! Novo formulário recebido ✅

Tipo de processo selecionado: {tipo_proc or "Não informado"}{sem_codigo_texto}
"""

    msg = MIMEMultipart()
    msg["From"] = smtp_user
    msg["To"] = smtp_user
    msg["Cc"] = ", ".join(cc_addr)
    msg["Subject"] = assunto
    msg.attach(MIMEText(corpo_email.strip(), "plain"))

    anexo = MIMEApplication(
        arquivo_bytes,
        _subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    nome_arquivo = st.session_state.nome_arquivo or "Pedido.xlsx"
    anexo.add_header('Content-Disposition', 'attachment', filename=nome_arquivo)
    msg.attach(anexo)

    if anexos:
        for arquivo in anexos:
            try:
                conteudo = arquivo.getvalue()
                anexo_extra = MIMEApplication(conteudo)
                anexo_extra.add_header(
                    'Content-Disposition',
                    'attachment',
                    filename=arquivo.name
                )
                msg.attach(anexo_extra)
            except Exception as e:
                print(f"Erro ao anexar arquivo extra {getattr(arquivo, 'name', 'sem_nome')}: {e}")

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
            print("📨 E-mail de pedido enviado com sucesso!")
    except Exception as e:
        print(f"Erro ao enviar e-mail: {e}")

def carregar_dados():
    """Carrega dados de empreendimentos e insumos."""
    df_empreend = pd.read_excel("Empreendimentos.xlsx")
    df_empreend.columns = df_empreend.columns.str.strip().str.upper()

    df_insumos = pd.read_excel("Insumos.xlsx")
    df_insumos["Min"] = pd.to_numeric(df_insumos.iloc[:, 3], errors="coerce")
    df_insumos["Max"] = pd.to_numeric(df_insumos.iloc[:, 4], errors="coerce")
    df_insumos["Basico"] = df_insumos["Min"].notna() & df_insumos["Max"].notna()
    df_insumos = df_insumos[df_insumos["Descrição"].notna() & (df_insumos["Descrição"].str.strip() != "")]

    df_empreend.loc[-1] = [""] * df_empreend.shape[1]
    df_empreend.index = df_empreend.index + 1
    df_empreend = df_empreend.sort_index()

    insumos_vazios = pd.DataFrame({"Código": [""], "Descrição": [""], "Unidade": [""]})
    df_insumos = pd.concat([insumos_vazios, df_insumos], ignore_index=True)
    df_insumos["Código"] = df_insumos["Código"].fillna("").astype(str)
    return df_empreend, df_insumos

# --- CARREGAMENTO DE DADOS ---
df_empreend, df_insumos = carregar_dados()

# --- LIMPEZA APÓS ENVIO ---
if st.session_state.get("limpar_pedido", False):
    for campo in ["pedido_numero", "solicitante", "executivo", "obra_selecionada", "cnpj", "endereco", "cep"]:
        if campo in st.session_state:
            try:
                st.session_state[campo] = ""
            except Exception:
                pass
    st.session_state.data_pedido = date.today()
    st.session_state.insumos = []
    st.session_state.limpar_pedido = False

# --- LOGO E CABEÇALHO ---
col1, col2, col3 = st.columns([1, 2, 1])
with col2:
    st.image("logo.png", width=300)

st.markdown("""
<div style='text-align: center;'>
    <h2 style='color: #000000;'>Pedido de Materiais</h2>
    <p style='font-size: 14px; color: #555;'>
        Preencha os campos com atenção. Verifique se todos os dados estão corretos antes de enviar.<br>
        Ao finalizar, o pedido será automaticamente enviado para o e-mail do setor de Suprimentos.<br>
        Você poderá baixar a planilha gerada após o envio, para registro ou controle.
    </p>
</div>
""", unsafe_allow_html=True)

TIPO_PEDIDO   = "Pedido de Materiais"
TIPO_ED       = "Criação de ED"

opcoes_tipo = [TIPO_PEDIDO, TIPO_ED]

valor_atual = st.session_state.get("tipo_processo", TIPO_PEDIDO)
if valor_atual not in opcoes_tipo:
    valor_atual = TIPO_PEDIDO

tipo_processo = st.radio(
    "Selecione o processo para este formulário:",
    options=opcoes_tipo,
    index=opcoes_tipo.index(valor_atual),
    horizontal=True
)

st.session_state["tipo_processo"] = tipo_processo

# --- DADOS DO PEDIDO ---
with st.expander("📋 Dados do Pedido", expanded=True):
    # --- RESET ---
    if st.session_state.resetar_pedido:
        st.session_state.pedido_numero = ""
        st.session_state.data_pedido = date.today()   # data do início do preenchimento
        st.session_state.solicitante = ""
        st.session_state.executivo = ""
        st.session_state.executivo_obra = ""
        st.session_state.exec_emails_obra = []
        st.session_state.executivo_manual = ""
        st.session_state.adm_obra = ""
        st.session_state.obra_selecionada = ""
        st.session_state.cnpj = ""
        st.session_state.endereco = ""
        st.session_state.cep = ""
        st.session_state.resetar_pedido = False

    # 🔒 Data fixa do pedido (fica a do dia em que começou a preencher; não muda em rerun)
    if "data_pedido" not in st.session_state or not st.session_state.data_pedido:
        st.session_state.data_pedido = date.today()

    # --- limpa executivo manual quando trocar de obra ---
    if "obra_anterior" not in st.session_state:
        st.session_state.obra_anterior = ""
    if st.session_state.get("obra_selecionada", "") != st.session_state.obra_anterior:
        st.session_state.executivo_manual = ""
        st.session_state.obra_anterior = st.session_state.get("obra_selecionada", "")

    # garante keys (caso rode direto sem reset)
    for k, v in {
        "executivo_manual": "",
        "executivo_obra": "",
        "executivo": "",
        "exec_emails_obra": []
    }.items():
        if k not in st.session_state:
            st.session_state[k] = v

    col1, col2 = st.columns(2)

    # --- COLUNA 1 ---
    with col1:
        st.text_input("Pedido Nº", key="pedido_numero")
        st.text_input("Solicitante", key="solicitante")
        st.selectbox(
            "Obra",
            df_empreend["EMPREENDIMENTO"].unique(),
            index=0,
            key="obra_selecionada"
        )

    obra_selecionada = st.session_state.get("obra_selecionada", "")
    is_obra_coringa = obra_selecionada in OBRAS_SEM_EXECUTIVO_FIXO

    # executivos pré-definidos por obra
    execs = OBRA_EXECUTIVOS.get(obra_selecionada, [])
    nomes_execs = [e.get("executivo", "").strip() for e in execs if e.get("executivo")]
    emails_execs = [e.get("email", "").strip() for e in execs if e.get("email")]

    # --- COLUNA 2 ---
    with col2:
        # ✅ Data fixa (somente leitura)
        st.text_input(
            "Data",
            value=st.session_state.data_pedido.strftime("%d/%m/%Y"),
            disabled=True
        )

        if is_obra_coringa:
            # ✅ para 9991/9992: escolhe executivo manualmente (a partir do OBRA_EXECUTIVOS)
            opcoes_executivo = [""] + EXECUTIVOS_OPCOES
            exec_manual = st.selectbox("Executivo", opcoes_executivo, index=0, key="executivo_manual")

            st.session_state.executivo = exec_manual
            st.session_state.executivo_obra = exec_manual

            # pega o email do executivo escolhido (procura dentro do OBRA_EXECUTIVOS)
            if exec_manual:
                email_manual = next(
                    (item.get("email", "") for lista in OBRA_EXECUTIVOS.values() for item in lista
                     if item.get("executivo") == exec_manual),
                    ""
                )
                st.session_state.exec_emails_obra = [email_manual] if (email_manual or "").strip() else []
            else:
                st.session_state.exec_emails_obra = []

        else:
            # ✅ obras normais: auto preenchimento (1 ou 2 executivos)
            if nomes_execs:
                st.session_state.executivo_obra = "; ".join(nomes_execs)  # exibição (1 ou 2)
                st.session_state.executivo = nomes_execs[0]              # usado no Excel/validação
                st.session_state.exec_emails_obra = emails_execs
            else:
                st.session_state.executivo_obra = ""
                st.session_state.executivo = ""
                st.session_state.exec_emails_obra = []

            st.text_input("Executivo", key="executivo_obra", disabled=True)

        # ADM segue manual (como você já faz)
        opcoes_adm = [""] + list(ADM_EMAILS.keys())
        st.selectbox("Administrativo da Obra", opcoes_adm, index=0, key="adm_obra")

    # --- autopreenche dados da obra (cnpj/endereco/cep) ---
    if obra_selecionada:
        filtro = df_empreend["EMPREENDIMENTO"] == obra_selecionada
        if filtro.any():
            dados_obra = df_empreend.loc[filtro].iloc[0]
            st.session_state.cnpj = dados_obra["CNPJ"]
            st.session_state.endereco = dados_obra["ENDERECO"]
            st.session_state.cep = dados_obra["CEP"]

    st.text_input("CNPJ/CPF", key="cnpj", disabled=True)
    st.text_input("Endereço", key="endereco", disabled=True)
    st.text_input("CEP", value=st.session_state.get("cep", ""), disabled=True)

# --- CAMPOS ESPECÍFICOS POR TIPO DE PROCESSO ---
anexos_processo = []  # default, para usar mais à frente

if st.session_state.tipo_processo == TIPO_ED:
    with st.expander("📄 Dados da ED / OF filha", expanded=True):
        st.text_input("Nº OF Mãe", key="num_of_mae")
        st.text_input("Fornecedor da OF filha", key="fornecedor_of_filha")
        anexos_processo = st.file_uploader(
            "Anexar documentos (planilhas, PDFs, prints, etc.)",
            type=["pdf", "xlsx", "xls", "csv", "png", "jpg", "jpeg"],
            accept_multiple_files=True,
            key="anexos_ed"
        )
        if anexos_processo:
            st.info(f"{len(anexos_processo)} arquivo(s) será(ão) enviado(s) junto com a requisição de ED.")

# --- ADIÇÃO DE INSUMOS ---
with st.expander("➕ Adicionar Insumo", expanded=True):

    if st.session_state.get("limpar_campos_insumo", False):
        # 🔹 Remove todos os valores dos campos
        for campo in ["descricao_exibicao", "descricao_livre", "codigo", "unidade", "quantidade", "complemento", "data_necessaria_txt"]:
            if campo in st.session_state:
                del st.session_state[campo]

        # 🔹 Garante valor padrão inicial
        st.session_state.quantidade = 1.0
        st.session_state.descricao_exibicao = ""
        st.session_state.complemento = ""
        st.session_state.data_necessaria_txt = ""
        st.session_state.limpar_campos_insumo = False
        st.rerun()  # 🔁 força recarregar já limpo
    
    df_insumos_lista = df_insumos.sort_values(by="Descrição", ascending=True).copy()
    df_insumos_lista["opcao_exibicao"] = df_insumos_lista.apply(
        lambda x: f"{x['Descrição']} – {x['Código']} ({x['Unidade']})" if pd.notna(x["Código"]) and str(x["Código"]).strip() != "" else x["Descrição"],
        axis=1
    )
    
    descricao_exibicao = st.selectbox(
        "Descrição do insumo",
        df_insumos_lista["opcao_exibicao"],
        key="descricao_exibicao"
    )

    dados_insumo = df_insumos_lista[df_insumos_lista["opcao_exibicao"] == descricao_exibicao].iloc[0]
    codigo_sel = str(dados_insumo["Código"]).strip()
    usando_base = codigo_sel != ""

    if usando_base:
        st.session_state.codigo = dados_insumo["Código"]
        st.session_state.unidade = dados_insumo["Unidade"]
        st.session_state.descricao = dados_insumo["Descrição"]
    else:
        st.session_state.codigo = ""
        st.session_state.descricao = ""

    if "unidade" not in st.session_state or not st.session_state.unidade:
        st.session_state.unidade = ""

    st.write("Ou preencha manualmente o Nome e Unidade se não estiver listado:")

    descricao_livre = st.text_input("Nome do insumo (livre)", key="descricao_livre", disabled=usando_base)
    st.text_input("Código do insumo", key="codigo", disabled=True)
    st.text_input("Unidade", key="unidade", disabled=usando_base)
    quantidade = st.number_input("Quantidade", min_value=0.0, value=float(st.session_state.get("quantidade", 1)), step=0.01, format="%g", key="quantidade")
    complemento = st.text_area(
        "Complemento, se necessário (Utilize para especificar medidas, marcas, cores e/ou tamanhos)",
        key="complemento"
    )
    data_necessaria_txt = st.text_input(
        "Data de necessidade do insumo",
        key="data_necessaria_txt",
        placeholder="DD/MM/YYYY"
    )
        
    if st.button("➕ Adicionar insumo"):
        descricao_final = st.session_state.descricao if usando_base else descricao_livre
    
        # valida data obrigatória somente na hora de adicionar
        dt, err = validar_data_br(st.session_state.get("data_necessaria_txt"))
    
        if err:
            st.warning(f"⚠️ {err}")
            st.stop()
    
        if descricao_final and quantidade > 0 and (usando_base or st.session_state.unidade.strip()):
            qtd = float(quantidade)
            if qtd.is_integer():
                qtd = int(qtd)
    
            novo_insumo = {
                "descricao": descricao_final,
                "codigo": st.session_state.codigo if usando_base else "",
                "unidade": st.session_state.unidade,
                "quantidade": qtd,
                "complemento": complemento,
                "data_necessaria": dt,  # <-- agora é date validado
            }
            st.session_state.insumos.append(novo_insumo)
    
            st.session_state.limpar_campos_insumo = True
            st.success("✅ Insumo adicionado com sucesso!")
            st.rerun()
        else:
            st.warning("⚠️ Preencha todos os campos obrigatórios do insumo.")
    
# --- Renderiza tabela de insumos ---
if st.session_state.insumos:
    st.markdown("""
    <style>
    /* Cabeçalhos */
    .tabela-header {
        font-weight: 600;
        color: #333;
        border-bottom: 2px solid #ccc;
        padding-bottom: 2px;
        margin-bottom: 2px;
        font-size: 15px;
        display: flex;
        align-items: center;
        justify-content: flex-start;
    }
    .tabela-header.center {
        justify-content: center; /* centraliza só Qtd */
    }

    /* Linhas */
    .linha-insumo {
        border-bottom: 1px solid #e6e6e6;
        padding: 3px 0;
        font-size: 14px;
        line-height: 1.4;
        display: flex;
        align-items: center;
    }

    .center {
        justify-content: center;
        text-align: center;
    }

    /* Botão 🗑️ */
    div[data-testid="stButton"] button {
        border: none;
        background-color: transparent;
        color: #666;
        font-size: 18px;
        padding: 0;
        line-height: 1;
        transform: translateY(-2px);
    }

    div[data-testid="stButton"] button:hover {
        color: #d9534f;
        transform: scale(1.15) translateY(-2px);
    }
    </style>
    """, unsafe_allow_html=True)

    # Cabeçalho
    col1, col2, col3, col4, col5 = st.columns([5.0, 1.0, 1.6, 1.0, 0.5])
    with col1:
        st.markdown("<div class='tabela-header'>Insumos Adicionados</div>", unsafe_allow_html=True)
    with col2:
        st.markdown("<div class='tabela-header center'>Qtd</div>", unsafe_allow_html=True)
    with col3:
        st.markdown("<div class='tabela-header center'>Entrega</div>", unsafe_allow_html=True)
    with col4:
        st.markdown("<div class='tabela-header'>Unid</div>", unsafe_allow_html=True)
    with col5:
        st.markdown("<div style='height:10px;'></div>", unsafe_allow_html=True)

    # Linhas
    for i, insumo in enumerate(st.session_state.insumos):
        col1, col2, col3, col4, col5 = st.columns([5.0, 1.0, 1.6, 1.0, 0.5])
    
        with col1:
            st.markdown(f"<div class='linha-insumo'>{insumo['descricao']}</div>", unsafe_allow_html=True)
    
        with col2:
            st.markdown(f"<div class='linha-insumo center'>{insumo['quantidade']}</div>", unsafe_allow_html=True)
    
        with col3:
            dt = insumo.get("data_necessaria")
            dt_txt = dt.strftime("%d/%m/%Y") if dt else ""
            st.markdown(f"<div class='linha-insumo center'>{dt_txt}</div>", unsafe_allow_html=True)
    
        with col4:
            st.markdown(f"<div class='linha-insumo'>{insumo['unidade']}</div>", unsafe_allow_html=True)
    
        with col5:
            if st.button("🗑️", key=f"delete_{i}"):
                st.session_state.insumos.pop(i)
                st.rerun()

# --- FINALIZAÇÃO DO PEDIDO ---
if st.button("📤 Enviar Pedido", use_container_width=True):

    tipo_proc = st.session_state.get("tipo_processo", TIPO_PEDIDO)
    
    campos_obrigatorios = [
        st.session_state.pedido_numero, st.session_state.data_pedido,
        st.session_state.solicitante, st.session_state.executivo,
        st.session_state.obra_selecionada, st.session_state.cnpj,
        st.session_state.endereco, st.session_state.cep, st.session_state.adm_obra
    ]
    if not st.session_state.adm_obra or st.session_state.adm_obra.strip() == "":
        st.warning("⚠️ Selecione o Administrativo da obra antes de enviar o pedido.")
        st.stop()
    if not all(campos_obrigatorios):
        st.warning("⚠️ Preencha todos os campos obrigatórios antes de enviar o pedido.")
        st.stop()
    if not st.session_state.insumos:
        st.warning("⚠️ Adicione pelo menos um insumo antes de enviar o pedido.")
        st.stop()

    if tipo_proc == TIPO_ED:
        num_of_mae = st.session_state.get("num_of_mae", "").strip()
        fornecedor_of_filha = st.session_state.get("fornecedor_of_filha", "").strip()

        if not num_of_mae:
            st.warning("⚠️ Informe o Nº da OF Mãe para criar a ED / OF filha.")
            st.stop()

        if not fornecedor_of_filha:
            st.warning("⚠️ Informe o Fornecedor da OF filha.")
            st.stop()

        if not anexos_processo:
            st.warning("⚠️ Anexe pelo menos um documento para a ED / OF filha.")
            st.stop()

    erro = None
    ok = False
    with st.spinner("Enviando pedido e gerando arquivo... Aguarde!"):
        try:
            caminho_modelo = "Modelo_Pedido.xlsx"
            wb = load_workbook(caminho_modelo)
            ws = wb["Pedido"]

            ws["F2"] = st.session_state.pedido_numero
            ws["C3"] = st.session_state.data_pedido.strftime("%d/%m/%Y")
            ws["C4"] = st.session_state.solicitante
            ws["C5"] = st.session_state.executivo
            ws["C7"] = st.session_state.obra_selecionada
            ws["C8"] = st.session_state.cnpj
            ws["C9"] = st.session_state.endereco
            ws["C10"] = st.session_state.cep

            linha = 13
            for insumo in st.session_state.insumos:
                ws[f"B{linha}"] = insumo["codigo"]
                ws[f"C{linha}"] = insumo["descricao"]
                ws[f"D{linha}"] = insumo["unidade"]
                ws[f"E{linha}"] = insumo["quantidade"]
                ws[f"F{linha}"] = insumo["complemento"]
                ws[f"G{linha}"] = (
                    insumo["data_necessaria"].strftime("%d/%m/%Y")
                    if insumo.get("data_necessaria") else ""
                )
                linha += 1

            ultima_linha_util = linha - 1
            ws.print_area = f"A1:G{ultima_linha_util}"

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            st.session_state.excel_bytes = buffer.read()
            st.session_state.nome_arquivo = f"Pedido{st.session_state.pedido_numero} OC {st.session_state.obra_selecionada}.xlsx"

            enviar_email_pedido(
                f"Pedido{st.session_state.pedido_numero} OC {st.session_state.obra_selecionada}",
                st.session_state.excel_bytes,
                st.session_state.insumos,
                ADM_EMAILS,
                anexos=anexos_processo
            )
            ok = True
        except Exception as e:
            erro = str(e)

    if ok:
        st.session_state.pedido_enviado = True
        st.success("✅ Pedido gerado e e-mail enviado com sucesso! Agora você pode baixar o arquivo Excel abaixo ⬇️")
    elif erro:
        st.error(f"❌ Erro ao gerar pedido: {erro}")

# --- BOTÕES APÓS ENVIO ---
if st.session_state.get("excel_bytes"):  # só renderiza se o arquivo existir
    col1, col2 = st.columns(2)

    with col1:
        if st.download_button(
            "📥 Baixar Excel",
            data=st.session_state.excel_bytes,
            file_name=st.session_state.nome_arquivo or "Pedido.xlsx",
            use_container_width=True
        ):
            # 🔹 Marca flags para limpar no próximo ciclo
            st.session_state.rerun_depois_download = True

    with col2:
        if st.button("🔄 Novo Pedido", use_container_width=True):
            for campo in [
                "pedido_numero", "solicitante", "executivo", "obra_selecionada",
                "cnpj", "endereco", "cep", "data_pedido",
                "excel_bytes", "nome_arquivo", "pedido_enviado",
                "adm_obra",
                "num_of_mae",
                "fornecedor_of_filha",
                "tipo_processo",
                "anexos_ed",
            ]:
                if campo in st.session_state:
                    del st.session_state[campo]
            
            st.session_state.insumos = []
            st.session_state["adm_obra"] = ""
            st.session_state["tipo_processo"] = TIPO_PEDIDO
            st.rerun()

# --- 🔄 KEEP-ALIVE (mantém app ativo no Streamlit Cloud) ---
st.components.v1.html("""
<script>
setInterval(() => {
    fetch(window.location.pathname + '_stcore/health');
}, 120000);
</script>
""", height=0)
